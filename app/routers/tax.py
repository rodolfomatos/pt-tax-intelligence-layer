"""
Tax analysis routers.

Contains endpoints for tax analysis, validation, batch processing, and export.
"""

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, Response
import csv
import io
from typing import Optional, List
from datetime import datetime

from app.models import (
    TaxAnalysisInput,
    TaxAnalysisOutput,
    TaxValidationInput,
    TaxValidationOutput,
)
from app.models.batch import BatchAnalysisRequest, BatchAnalysisResponse
from app.services.rules.engine import get_rule_engine
from app.services.reasoning import get_llm_reasoning
from app.services.decision import get_decision_aggregator
from app.data.ptdata.client import get_ptdata_client
from app.data.cache.client import get_cache_client
from app.database.audit import get_audit_repository
from app.services.hooks import get_system_hooks, EventType
import logging
import time

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/tax", tags=["Tax Analysis"])


@router.post("/analyze", response_model=TaxAnalysisOutput)
async def analyze_tax(input: TaxAnalysisInput, request: Request):
    """
    Main tax analysis endpoint.

    Processes a tax analysis request through the decision pipeline:
    1. Rule engine (deterministic) - tries to match against known rules
    2. LLM reasoning - if no clear rule match, uses AI for analysis
    3. Decision aggregator - combines results and produces final decision

    All decisions are logged for auditability and stored in semantic memory.
    """
    start_time = time.time()
    request_id = request.headers.get("X-Request-ID", "unknown")

    logger.info(f"Analyzing: {input.operation_type} - {input.description}")

    rule_engine = get_rule_engine()
    llm_reasoning = get_llm_reasoning()
    aggregator = get_decision_aggregator()

    # Step 1: Try rule engine (deterministic)
    rule_result = rule_engine.evaluate(input)
    source = "rule_engine" if rule_result else "llm"

    # Step 2: Try LLM reasoning (if no clear rule)
    llm_result = await llm_reasoning.analyze(input)

    # Step 3: Aggregate and decide
    final_result = await aggregator.decide(input, rule_result, llm_result)

    processing_time_ms = int((time.time() - start_time) * 1000)

    # Log to audit database
    try:
        audit = get_audit_repository()
        await audit.log_decision(
            input_data=input,
            output_data=final_result,
            source=source,
            processing_time_ms=processing_time_ms,
        )
    except Exception as e:
        logger.error(f"Failed to log decision: {e}")

    # Log action
    try:
        audit = get_audit_repository()
        await audit.log_action(
            action="tax_analyze",
            entity_type=input.entity_type,
            user=request.headers.get("X-User", "anonymous"),
            request_id=request_id,
            details={
                "operation_type": input.operation_type,
                "decision": final_result.decision,
                "confidence": final_result.confidence,
            },
            ip_address=request.client.host if request.client else None,
        )
    except Exception as e:
        logger.error(f"Failed to log action: {e}")

    # Save to semantic memory (L3 layer)
    try:
        from app.data.memory.layers import get_memory_layers
        from app.data.memory.hooks import DecisionHooks

        memory = get_memory_layers()
        memory.save_to_memory(
            decision_id=str(id(final_result)),
            description=input.description,
            decision=final_result.decision,
            explanation=final_result.explanation,
            legal_basis=[lb.model_dump() for lb in final_result.legal_basis],
            metadata={
                "operation_type": input.operation_type,
                "entity_type": input.entity_type,
                "project_type": input.context.project_type,
                "source": source,
            },
        )

        DecisionHooks.on_decision(
            decision_id=str(id(final_result)),
            input_data=input.model_dump(),
            output_data=final_result.model_dump(),
        )
    except Exception as e:
        logger.error(f"Failed to save to memory: {e}")

    # Save to knowledge graph (GMIF classification)
    try:
        from app.data.memory.graph.builder import get_graph_builder
        from app.data.memory.graph.gmif import get_gmif_classifier

        builder = get_graph_builder()
        gmif_type = await builder.add_decision(
            decision_id=str(id(final_result)),
            description=input.description,
            decision_type=final_result.decision,
            confidence=final_result.confidence,
            legal_basis=[lb.model_dump() for lb in final_result.legal_basis],
            entity_type=input.entity_type,
            project_type=input.context.project_type,
            risks=final_result.risks,
            assumptions=final_result.assumptions,
        )
        logger.info(f"GMIF classification: {gmif_type}")
    except Exception as e:
        logger.warning(f"Failed to save to knowledge graph: {e}")

    logger.info(
        f"Decision: {final_result.decision} (confidence: {final_result.confidence})"
    )
    return final_result


@router.post("/validate", response_model=TaxValidationOutput)
async def validate_tax(input: TaxValidationInput, request: Request):
    """
    Validate an existing decision for consistency.

    Performs consistency checks on a decision:
    - High confidence but limited legal basis → warning
    - Uncertain decision with high confidence → warning
    - High risk with low confidence → note
    - Missing legal citation fields → warning
    """
    notes = []
    warnings = []

    if input.confidence > 0.8 and len(input.legal_basis) < 2:
        warnings.append("High confidence but limited legal basis")

    if input.decision == "uncertain" and input.confidence > 0.7:
        warnings.append("Uncertain decision should have lower confidence")

    if input.risk_level == "high" and input.confidence < 0.5:
        notes.append("High risk with low confidence - review recommended")

    for citation in input.legal_basis:
        if not citation.code or not citation.article:
            warnings.append("Legal citation missing code or article")

    return TaxValidationOutput(
        valid=len(warnings) == 0,
        consistency_check="passed" if len(warnings) == 0 else "warnings",
        notes=notes,
        warnings=warnings,
    )


@router.post("/analyze/batch", response_model=BatchAnalysisResponse)
async def analyze_tax_batch(request: Request, batch: BatchAnalysisRequest):
    """
    Batch tax analysis endpoint.

    Processes multiple tax analysis requests in a single call.
    Useful for bulk operations (e.g., university expense reports).
    """
    results = []
    errors = []
    successful = 0
    failed = 0

    for i, input_item in enumerate(batch.items):
        try:
            rule_engine = get_rule_engine()
            llm_reasoning = get_llm_reasoning()
            aggregator = get_decision_aggregator()

            rule_result = rule_engine.evaluate(input_item)
            source = "rule_engine" if rule_result else "llm"

            llm_result = await llm_reasoning.analyze(input_item)
            final_result = await aggregator.decide(input_item, rule_result, llm_result)

            results.append(final_result)
            successful += 1

            # Trigger webhook for high-risk decisions
            if final_result.risk_level == "high":
                try:
                    hooks = get_system_hooks()
                    await hooks.trigger(
                        EventType.DECISION_HIGH_RISK,
                        payload=final_result.model_dump(),
                    )
                except Exception as e:
                    logger.warning(f"Failed to trigger high-risk webhook: {e}")

            # Log to audit
            try:
                audit = get_audit_repository()
                await audit.log_decision(
                    input_data=input_item,
                    output_data=final_result,
                    source=source,
                )
            except Exception as e:
                logger.warning(f"Failed to log batch decision to audit: {e}")

        except Exception as e:
            failed += 1
            error_detail = {
                "index": i,
                "error": str(e),
                "input": input_item.model_dump(),
            }
            errors.append(error_detail)
            logger.error(f"Batch item {i} failed: {e}")

            if batch.stop_on_error:
                break

    logger.info(f"Batch complete: {successful} success, {failed} failed")

    return BatchAnalysisResponse(
        total=len(batch.items),
        successful=successful,
        failed=failed,
        results=results,
        errors=errors,
    )


@router.get("/export")
async def export_decisions(
    format: str = "csv",
    decision_type: Optional[str] = None,
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """
    Export tax decisions to CSV or Excel.

    Generates a file with all decisions matching the filters.
    Useful for university reporting and audit purposes.
    """
    try:
        from fastapi.responses import Response as FastAPIResponse
        import csv
        import io

        audit = get_audit_repository()
        decisions = await audit.get_decisions(
            limit=10000,
            offset=0,
            decision_type=decision_type,
            entity_type=entity_type,
        )

        if format == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tax Decisions"

                # Headers
                headers = [
                    "ID",
                    "Date",
                    "Operation",
                    "Description",
                    "Amount",
                    "Decision",
                    "Confidence",
                    "Risk",
                    "Source",
                ]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

                # Data
                for row, d in enumerate(decisions, 2):
                    ws.cell(row=row, column=1, value=str(d.id))
                    ws.cell(row=row, column=2, value=d.created_at.isoformat())
                    ws.cell(row=row, column=3, value=d.operation_type)
                    ws.cell(row=row, column=4, value=d.description)
                    ws.cell(row=row, column=5, value=float(d.amount))
                    ws.cell(row=row, column=6, value=d.decision)
                    ws.cell(row=row, column=7, value=float(d.confidence))
                    ws.cell(row=row, column=8, value=d.risk_level)
                    ws.cell(row=row, column=9, value=d.source)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                return Response(
                    content=output.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": "attachment; filename=tax_decisions.xlsx"
                    },
                )
            except ImportError:
                logger.warning("openpyxl not installed, falling back to CSV")
                format = "csv"

        # CSV export
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "ID",
                "Date",
                "Operation",
                "Description",
                "Amount",
                "Decision",
                "Confidence",
                "Risk",
                "Source",
            ]
        )

        for d in decisions:
            writer.writerow(
                [
                    str(d.id),
                    d.created_at.isoformat(),
                    d.operation_type,
                    d.description,
                    float(d.amount),
                    d.decision,
                    float(d.confidence),
                    d.risk_level,
                    d.source,
                ]
            )

        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=tax_decisions.csv"},
        )

    except Exception as e:
        logger.error(f"Export failed: {e}")
        raise HTTPException(status_code=500, detail="Export failed")
